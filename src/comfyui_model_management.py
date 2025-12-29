import os
import sys
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSize, QTimer, QSettings
from PyQt6.QtGui import QColor, QPalette
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QHBoxLayout,
    QVBoxLayout,
    QLineEdit,
    QPushButton,
    QFileDialog,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QMessageBox,
    QProgressBar,
    QLabel,
    QStyle,
    QCheckBox,
    QComboBox,
    QAbstractItemView,
)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

MODEL_EXTS = {".safetensors", ".ckpt", ".pth", ".pt", ".onnx", ".bin", ".gguf"}


# ---------------- Utilities ----------------

def fmt_dt(ts: float) -> str:
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def parse_dt(s: str) -> Optional[datetime]:
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def fmt_bytes(n: int) -> str:
    n = int(n)
    if n < 1024:
        return f"{n} B"
    kb = n / 1024
    if kb < 1024:
        return f"{kb:.1f} KB"
    mb = kb / 1024
    if mb < 1024:
        return f"{mb:.1f} MB"
    gb = mb / 1024
    if gb < 1024:
        return f"{gb:.2f} GB"
    tb = gb / 1024
    return f"{tb:.2f} TB"


# ---------------- Data ----------------

@dataclass
class FileRow:
    full_path: str        # internal use only
    directory: str
    name: str
    length: int
    last_access_time: str
    last_write_time: str
    creation_time: str


# ---------------- Workers ----------------

class ScanWorker(QThread):
    """
    Streaming scan:
    - row_found emits each FileRow as soon as it is read
    - progress emits 0..100 based on processed candidates
    """
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    row_found = pyqtSignal(object)   # FileRow
    done = pyqtSignal(int)           # total
    error = pyqtSignal(str)

    def __init__(self, root_dir: str, parent=None):
        super().__init__(parent)
        self.root_dir = root_dir
        self._stop = False

    def request_stop(self):
        self._stop = True

    def run(self):
        try:
            root = Path(self.root_dir)
            if not root.exists() or not root.is_dir():
                self.error.emit("Please select a valid ComfyUI directory.")
                return

            self.status.emit("Scanning: counting candidate files...")
            candidates: List[Path] = []
            for dirpath, _, filenames in os.walk(root):
                if self._stop:
                    return
                for fn in filenames:
                    if self._stop:
                        return
                    p = Path(dirpath) / fn
                    if p.suffix.lower() in MODEL_EXTS:
                        candidates.append(p)

            total = len(candidates)
            if total == 0:
                self.progress.emit(100)
                self.status.emit("Scan complete: no model files found.")
                self.done.emit(0)
                return

            self.status.emit(f"Scanning: reading metadata for {total} files...")
            for i, p in enumerate(candidates, start=1):
                if self._stop:
                    return
                try:
                    st = p.stat()
                    row = FileRow(
                        full_path=str(p),
                        directory=str(p.parent),
                        name=p.name,
                        length=int(st.st_size),
                        last_access_time=fmt_dt(st.st_atime),
                        last_write_time=fmt_dt(st.st_mtime),
                        creation_time=fmt_dt(st.st_ctime),
                    )
                    self.row_found.emit(row)
                except Exception:
                    pass

                self.progress.emit(int((i / total) * 100))

            self.status.emit(f"Scan complete: found {total} files.")
            self.done.emit(total)

        except Exception as e:
            self.error.emit(f"Scan failed: {e}\n\n{traceback.format_exc()}")


class DeleteWorker(QThread):
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    deleted = pyqtSignal(list)  # paths
    failed = pyqtSignal(list)   # (path, err)
    error = pyqtSignal(str)

    def __init__(self, paths: List[str], use_recycle_bin: bool, parent=None):
        super().__init__(parent)
        self.paths = paths
        self.use_recycle_bin = use_recycle_bin
        self._stop = False

    def request_stop(self):
        self._stop = True

    def run(self):
        try:
            total = len(self.paths)
            if total == 0:
                self.progress.emit(0)
                self.status.emit("Nothing selected to delete.")
                self.deleted.emit([])
                self.failed.emit([])
                return

            deleter = self._delete_to_recycle_bin if self.use_recycle_bin else self._delete_permanently
            mode = "Recycle Bin" if self.use_recycle_bin else "permanent delete"
            self.status.emit(f"Deleting {total} files ({mode})...")

            deleted: List[str] = []
            failed: List[Tuple[str, str]] = []

            for i, p in enumerate(self.paths, start=1):
                if self._stop:
                    return
                try:
                    deleter(p)
                    deleted.append(p)
                except Exception as e:
                    failed.append((p, str(e)))

                self.progress.emit(int((i / total) * 100))

            self.status.emit(f"Delete finished: {len(deleted)} deleted, {len(failed)} failed.")
            self.deleted.emit(deleted)
            self.failed.emit(failed)

        except Exception as e:
            self.error.emit(f"Delete failed: {e}\n\n{traceback.format_exc()}")

    @staticmethod
    def _delete_permanently(path: str):
        os.remove(path)

    @staticmethod
    def _delete_to_recycle_bin(path: str):
        try:
            from send2trash import send2trash  # type: ignore
        except Exception as e:
            raise RuntimeError(
                "Recycle Bin delete requires the 'send2trash' package.\n"
                "Install it with: pip install send2trash\n"
                f"Original import error: {e}"
            )
        send2trash(path)


# ---------------- Frozen Grid ----------------

class FrozenGrid(QWidget):
    """
    Left table: checkbox column frozen (always visible)
    Right table: data columns horizontally scrollable, resizable, sortable by clicking headers (handled by MainWindow)
    """
    checkbox_toggled = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)

        self.left = QTableWidget(0, 1)
        self.right = QTableWidget(0, 6)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addWidget(self.left)
        layout.addWidget(self.right, 1)

        self._setup_tables()
        self._wire_sync()

    def _setup_tables(self):
        # Left (checkbox)
        self.left.setHorizontalHeaderLabels([""])
        self.left.verticalHeader().setVisible(False)
        self.left.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.left.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.left.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.left.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.left.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        lh = self.left.horizontalHeader()
        lh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.left.setColumnWidth(0, 30)

        # Make left table tight (no wasted space)
        fw = self.left.frameWidth()
        tight = self.left.columnWidth(0) + (fw * 2) + 2
        self.left.setMinimumWidth(tight)
        self.left.setMaximumWidth(tight)

        # Right (data)
        self.right.setHorizontalHeaderLabels(
            ["Directory", "Name", "Length", "LastAccessTime", "LastWriteTime", "CreationTime"]
        )
        self.right.verticalHeader().setVisible(False)
        self.right.setAlternatingRowColors(True)
        self.right.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.right.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.right.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.right.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        # Allow column resizing with mouse
        rh = self.right.horizontalHeader()
        rh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        rh.setStretchLastSection(False)

        # initial widths (user can resize)
        self.right.setColumnWidth(0, 520)
        self.right.setColumnWidth(1, 260)
        self.right.setColumnWidth(2, 120)
        self.right.setColumnWidth(3, 175)
        self.right.setColumnWidth(4, 175)
        self.right.setColumnWidth(5, 175)

    def _wire_sync(self):
        # sync vertical scroll
        self.left.verticalScrollBar().valueChanged.connect(self.right.verticalScrollBar().setValue)
        self.right.verticalScrollBar().valueChanged.connect(self.left.verticalScrollBar().setValue)

        # sync row heights
        self.left.verticalHeader().sectionResized.connect(self._sync_row_height_from_left)
        self.right.verticalHeader().sectionResized.connect(self._sync_row_height_from_right)

        # checkbox toggles
        self.left.itemChanged.connect(self._on_left_item_changed)
        self.left.cellPressed.connect(self._on_left_cell_pressed)

        # keep left scrolled with current selection on right
        self.right.currentCellChanged.connect(self._sync_current_row)

    def _sync_row_height_from_left(self, row: int, old: int, new: int):
        if self.right.rowCount() > row and self.right.rowHeight(row) != new:
            self.right.setRowHeight(row, new)

    def _sync_row_height_from_right(self, row: int, old: int, new: int):
        if self.left.rowCount() > row and self.left.rowHeight(row) != new:
            self.left.setRowHeight(row, new)

    def _sync_current_row(self, currentRow: int, currentCol: int, prevRow: int, prevCol: int):
        if currentRow >= 0:
            it = self.left.item(currentRow, 0)
            if it:
                self.left.scrollToItem(it, QAbstractItemView.ScrollHint.PositionAtCenter)

    def _on_left_cell_pressed(self, row: int, col: int):
        if col == 0:
            QTimer.singleShot(0, self.checkbox_toggled.emit)

    def _on_left_item_changed(self, item: QTableWidgetItem):
        if item.column() == 0:
            self.checkbox_toggled.emit()

    def set_sort_indicator(self, col: int, ascending: bool):
        arrow = "▲" if ascending else "▼"
        base = ["Directory", "Name", "Length", "LastAccessTime", "LastWriteTime", "CreationTime"]
        base = [h.replace(" ▲", "").replace(" ▼", "") for h in base]
        if 0 <= col < len(base):
            base[col] = f"{base[col]} {arrow}"
        self.right.setHorizontalHeaderLabels(base)

    def clear(self):
        self.left.blockSignals(True)
        self.right.blockSignals(True)
        self.left.setRowCount(0)
        self.right.setRowCount(0)
        self.left.blockSignals(False)
        self.right.blockSignals(False)

    def row_count(self) -> int:
        return self.right.rowCount()

    def populate(self, rows: List[FileRow], checked_paths: set[str]):
        self.left.blockSignals(True)
        self.right.blockSignals(True)

        self.left.setRowCount(0)
        self.right.setRowCount(0)

        for r in rows:
            row_idx = self.right.rowCount()
            self.left.insertRow(row_idx)
            self.right.insertRow(row_idx)

            chk = QTableWidgetItem()
            chk.setFlags(
                Qt.ItemFlag.ItemIsEnabled |
                Qt.ItemFlag.ItemIsUserCheckable |
                Qt.ItemFlag.ItemIsSelectable
            )
            chk.setData(Qt.ItemDataRole.UserRole, r.full_path)
            chk.setCheckState(Qt.CheckState.Checked if r.full_path in checked_paths else Qt.CheckState.Unchecked)
            self.left.setItem(row_idx, 0, chk)

            def set_cell(col: int, text: str, align_right: bool = False):
                it = QTableWidgetItem(text)
                if align_right:
                    it.setTextAlignment(int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter))
                self.right.setItem(row_idx, col, it)

            set_cell(0, r.directory)
            set_cell(1, r.name)
            set_cell(2, str(r.length), align_right=True)
            set_cell(3, r.last_access_time)
            set_cell(4, r.last_write_time)
            set_cell(5, r.creation_time)

            h = self.right.rowHeight(row_idx)
            self.left.setRowHeight(row_idx, h)

        self.left.blockSignals(False)
        self.right.blockSignals(False)

    def append_row(self, r: FileRow, is_checked: bool = False):
        """
        Used for streaming during scan (no sorting here).
        """
        self.left.blockSignals(True)
        self.right.blockSignals(True)

        row_idx = self.right.rowCount()
        self.left.insertRow(row_idx)
        self.right.insertRow(row_idx)

        chk = QTableWidgetItem()
        chk.setFlags(
            Qt.ItemFlag.ItemIsEnabled |
            Qt.ItemFlag.ItemIsUserCheckable |
            Qt.ItemFlag.ItemIsSelectable
        )
        chk.setData(Qt.ItemDataRole.UserRole, r.full_path)
        chk.setCheckState(Qt.CheckState.Checked if is_checked else Qt.CheckState.Unchecked)
        self.left.setItem(row_idx, 0, chk)

        self.right.setItem(row_idx, 0, QTableWidgetItem(r.directory))
        self.right.setItem(row_idx, 1, QTableWidgetItem(r.name))

        it_len = QTableWidgetItem(str(r.length))
        it_len.setTextAlignment(int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter))
        self.right.setItem(row_idx, 2, it_len)

        self.right.setItem(row_idx, 3, QTableWidgetItem(r.last_access_time))
        self.right.setItem(row_idx, 4, QTableWidgetItem(r.last_write_time))
        self.right.setItem(row_idx, 5, QTableWidgetItem(r.creation_time))

        h = self.right.rowHeight(row_idx)
        self.left.setRowHeight(row_idx, h)

        self.left.blockSignals(False)
        self.right.blockSignals(False)

    def selected_paths(self) -> List[str]:
        paths = []
        for i in range(self.left.rowCount()):
            it = self.left.item(i, 0)
            if it and it.checkState() == Qt.CheckState.Checked:
                p = it.data(Qt.ItemDataRole.UserRole)
                if p:
                    paths.append(str(p))
        return paths

    def any_checked(self) -> bool:
        for i in range(self.left.rowCount()):
            it = self.left.item(i, 0)
            if it and it.checkState() == Qt.CheckState.Checked:
                return True
        return False

    def selected_count_and_size(self) -> Tuple[int, int]:
        count = 0
        total = 0
        for i in range(self.right.rowCount()):
            chk = self.left.item(i, 0)
            if chk and chk.checkState() == Qt.CheckState.Checked:
                count += 1
                length_item = self.right.item(i, 2)
                if length_item:
                    try:
                        total += int(length_item.text())
                    except Exception:
                        pass
        return count, total

    def visible_rows_as_filerows(self) -> List[FileRow]:
        out: List[FileRow] = []
        for i in range(self.right.rowCount()):
            chk = self.left.item(i, 0)
            full_path = str(chk.data(Qt.ItemDataRole.UserRole)) if chk else ""
            directory = self.right.item(i, 0).text()
            name = self.right.item(i, 1).text()
            length = int(self.right.item(i, 2).text() or "0")
            atime = self.right.item(i, 3).text()
            mtime = self.right.item(i, 4).text()
            ctime = self.right.item(i, 5).text()
            out.append(FileRow(full_path, directory, name, length, atime, mtime, ctime))
        return out


# ---------------- Main Window ----------------

class MainWindow(QMainWindow):
    MODE_IDLE = "idle"
    MODE_SCANNING = "scanning"
    MODE_DELETING = "deleting"

    # right table column indexes
    SORT_DIRECTORY = 0
    SORT_NAME = 1
    SORT_LENGTH = 2
    SORT_ATIME = 3
    SORT_MTIME = 4
    SORT_CTIME = 5

    def __init__(self):
        super().__init__()
        self.setWindowTitle("ComfyUI Model Scanner")
        self.setMinimumSize(QSize(1200, 700))

        self.scan_worker: Optional[ScanWorker] = None
        self.del_worker: Optional[DeleteWorker] = None

        self.all_rows: List[FileRow] = []
        self.progress_mode = self.MODE_IDLE

        # default sort: LastAccessTime ascending (least accessed first)
        self.sort_col = self.SORT_ATIME
        self.sort_ascending = True

        # streaming UI batching
        self._stream_buffer: List[FileRow] = []
        self._stream_timer = QTimer(self)
        self._stream_timer.setInterval(150)
        self._stream_timer.timeout.connect(self._flush_stream_buffer)

        self._build_ui()
        self._wire_events()
        self._apply_polish()

        self._load_settings()
        self._apply_sort_indicator()
        self._set_progress_mode(self.MODE_IDLE)
        self._refresh_action_states()

    # ---------- UI ----------

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)

        root = QHBoxLayout(central)

        left = QVBoxLayout()
        root.addLayout(left, 1)

        dir_row = QHBoxLayout()
        left.addLayout(dir_row)
        dir_row.addWidget(QLabel("ComfyUI Directory:"))

        self.dir_edit = QLineEdit()
        self.dir_edit.setPlaceholderText("Select your ComfyUI directory (e.g. D:\\ComfyUI)")
        dir_row.addWidget(self.dir_edit, 1)

        self.browse_btn = QPushButton("Browse…")
        dir_row.addWidget(self.browse_btn)

        filter_row = QHBoxLayout()
        left.addLayout(filter_row)
        filter_row.addWidget(QLabel("Filter:"))

        self.filter_edit = QLineEdit()
        self.filter_edit.setPlaceholderText("Type to filter by name / path / extension (e.g. 'vae' or '.safetensors')")
        filter_row.addWidget(self.filter_edit, 1)

        self.clear_filter_btn = QPushButton("Clear")
        filter_row.addWidget(self.clear_filter_btn)

        self.grid = FrozenGrid()
        left.addWidget(self.grid, 1)

        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setTextVisible(True)
        self.progress.setAlignment(Qt.AlignmentFlag.AlignCenter)
        left.addWidget(self.progress)

        self.status_label = QLabel("Ready.")
        left.addWidget(self.status_label)

        right = QVBoxLayout()
        root.addLayout(right)

        theme_row = QHBoxLayout()
        theme_row.addWidget(QLabel("Theme:"))
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["Light", "Dark"])
        theme_row.addWidget(self.theme_combo)
        right.addLayout(theme_row)

        right.addSpacing(8)

        self.scan_btn = QPushButton("Scan")
        self.select_all_btn = QPushButton("Select All")
        self.select_none_btn = QPushButton("Select None")
        self.delete_btn = QPushButton("Delete")
        self.download_btn = QPushButton("Download")

        for b in (self.scan_btn, self.select_all_btn, self.select_none_btn, self.delete_btn, self.download_btn):
            b.setMinimumWidth(170)

        right.addWidget(self.scan_btn)
        right.addSpacing(6)
        right.addWidget(self.select_all_btn)
        right.addWidget(self.select_none_btn)
        right.addSpacing(10)

        self.recycle_chk = QCheckBox("Move to Recycle Bin")
        self.recycle_chk.setChecked(True)
        right.addWidget(self.recycle_chk)

        tip = QLabel("Tip: Uncheck for permanent delete.")
        tip.setWordWrap(True)
        right.addWidget(tip)

        right.addSpacing(10)
        right.addWidget(self.delete_btn)
        right.addWidget(self.download_btn)
        right.addStretch(1)

        # initial state
        self.delete_btn.setEnabled(False)
        self.download_btn.setEnabled(False)
        self.select_all_btn.setEnabled(False)
        self.select_none_btn.setEnabled(False)
        self.scan_btn.setEnabled(False)  # enabled once folder valid

    def _wire_events(self):
        self.browse_btn.clicked.connect(self.on_browse)
        self.scan_btn.clicked.connect(self.on_scan)
        self.delete_btn.clicked.connect(self.on_delete)
        self.download_btn.clicked.connect(self.on_download)

        self.select_all_btn.clicked.connect(self.on_select_all)
        self.select_none_btn.clicked.connect(self.on_select_none)

        self.clear_filter_btn.clicked.connect(lambda: self.filter_edit.setText(""))
        self.filter_edit.textChanged.connect(lambda: self.apply_filter_and_refresh(force=False))

        self.dir_edit.textChanged.connect(self._validate_dir_and_update_scan_button)

        self.grid.checkbox_toggled.connect(self._refresh_action_states)

        # sorting by clicking headers
        self.grid.right.horizontalHeader().sectionClicked.connect(self.on_right_header_clicked)

        # theme
        self.theme_combo.currentTextChanged.connect(self.on_theme_changed)

    def _apply_polish(self):
        self.scan_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_BrowserReload))
        self.delete_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_TrashIcon))
        self.download_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))

    # ---------- Theme ----------

    def _apply_app_stylesheet(self, theme: str):
        if theme.lower() != "dark":
            QApplication.instance().setStyleSheet("""
            QPushButton:disabled { color: #777; }
            """)
            return

        QApplication.instance().setStyleSheet("""
        /* --- Table checkbox indicator visibility --- */
        QTableView::indicator {
            width: 16px;
            height: 16px;
        }
        QTableView::indicator:unchecked {
            border: 1px solid #DADADA;
            background: #2B2B2B;
        }
        QTableView::indicator:checked {
            border: 1px solid #DADADA;
            background: #3B82F6;
        }

        /* --- High contrast buttons --- */
        QPushButton {
            color: #F2F2F2;
            background-color: #3A3A3A;
            border: 1px solid #5A5A5A;
            border-radius: 6px;
            padding: 6px 10px;
        }
        QPushButton:hover { background-color: #444444; border: 1px solid #6A6A6A; }
        QPushButton:pressed { background-color: #2F2F2F; }
        QPushButton:disabled {
            color: #9A9A9A;
            background-color: #2E2E2E;
            border: 1px solid #3A3A3A;
        }

        /* Make checkbox + combo text more readable too */
        QCheckBox, QLabel {
            color: #F2F2F2;
        }

        /* ComboBox (control) */
        QComboBox {
            color: #F2F2F2;
            background-color: #2B2B2B;
            border: 1px solid #5A5A5A;
            border-radius: 6px;
            padding: 4px 8px;
        }
        QComboBox:disabled {
            color: #9A9A9A;
            background-color: #242424;
            border: 1px solid #3A3A3A;
        }
        QComboBox::drop-down {
            subcontrol-origin: padding;
            subcontrol-position: top right;
            width: 24px;
            border-left: 1px solid #5A5A5A;
        }

        /* ComboBox popup list (prevents faded items) */
        QComboBox QAbstractItemView {
            color: #F2F2F2;
            background-color: #1F1F1F;
            border: 1px solid #5A5A5A;
            selection-background-color: #3B82F6;
            selection-color: #FFFFFF;
            outline: 0;
        }
        QComboBox QAbstractItemView::item {
            padding: 6px 10px;
            color: #F2F2F2;
            background: transparent;
        }
        QComboBox QAbstractItemView::item:selected {
            color: #FFFFFF;
            background: #3B82F6;
        }

        QLineEdit {
            color: #F2F2F2;
            background-color: #242424;
            border: 1px solid #5A5A5A;
            border-radius: 6px;
            padding: 4px 8px;
        }
        """)

    def on_theme_changed(self, theme: str):
        self.apply_theme(theme)
        self._save_settings()

    def apply_theme(self, theme: str):
        app = QApplication.instance()
        if theme.lower() == "dark":
            palette = QPalette()
            palette.setColor(QPalette.ColorRole.Window, QColor(30, 30, 30))
            palette.setColor(QPalette.ColorRole.WindowText, QColor(230, 230, 230))
            palette.setColor(QPalette.ColorRole.Base, QColor(22, 22, 22))
            palette.setColor(QPalette.ColorRole.AlternateBase, QColor(30, 30, 30))
            palette.setColor(QPalette.ColorRole.Text, QColor(230, 230, 230))
            palette.setColor(QPalette.ColorRole.Button, QColor(45, 45, 45))
            palette.setColor(QPalette.ColorRole.ButtonText, QColor(240, 240, 240))
            palette.setColor(QPalette.ColorRole.Highlight, QColor(70, 120, 200))
            palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
            app.setPalette(palette)
        else:
            app.setPalette(app.style().standardPalette())

        self._apply_app_stylesheet(theme)

    # ---------- Settings ----------

    def _settings(self) -> QSettings:
        ini_path = str(Path(__file__).with_name("settings.ini"))
        return QSettings(ini_path, QSettings.Format.IniFormat)

    def _load_settings(self):
        s = self._settings()

        folder = s.value("comfyui_folder", "", type=str)
        if folder:
            self.dir_edit.setText(folder)

        theme = s.value("theme", "Light", type=str)
        if theme not in ("Light", "Dark"):
            theme = "Light"
        self.theme_combo.setCurrentText(theme)
        self.apply_theme(theme)

        self.sort_col = int(s.value("sort_col", self.SORT_ATIME))
        self.sort_ascending = bool(int(s.value("sort_ascending", 1)))
        self._apply_sort_indicator()

        self._validate_dir_and_update_scan_button()

    def _save_settings(self):
        s = self._settings()
        s.setValue("comfyui_folder", self.dir_edit.text().strip())
        s.setValue("theme", self.theme_combo.currentText())
        s.setValue("sort_col", int(self.sort_col))
        s.setValue("sort_ascending", 1 if self.sort_ascending else 0)
        s.sync()

    def closeEvent(self, event):
        self._save_settings()
        super().closeEvent(event)

    # ---------- Progress mode ----------

    def _set_progress_mode(self, mode: str):
        self.progress_mode = mode
        if mode == self.MODE_SCANNING:
            self.progress.setValue(0)
            self.progress.setFormat("Scanning… %p%")
        elif mode == self.MODE_DELETING:
            self.progress.setValue(0)
            self.progress.setFormat("Deleting… %p%")
        else:
            self._update_progress_summary()
        self._refresh_action_states()

    def _update_progress_summary(self):
        count, total_bytes = self.grid.selected_count_and_size()
        self.progress.setValue(100)  # reliable repaint for text
        if count == 0:
            self.progress.setFormat("Selected: 0 files")
        else:
            self.progress.setFormat(f"Selected: {count} files, {fmt_bytes(total_bytes)}")
        self.progress.repaint()

    # ---------- State helpers ----------

    def _validate_dir_and_update_scan_button(self):
        if self.progress_mode != self.MODE_IDLE:
            self.scan_btn.setEnabled(False)
            return
        p = self.dir_edit.text().strip()
        ok = bool(p) and Path(p).exists() and Path(p).is_dir()
        self.scan_btn.setEnabled(ok)

    def _refresh_action_states(self):
        idle = (self.progress_mode == self.MODE_IDLE)
        has_rows = self.grid.row_count() > 0
        has_selection = self.grid.any_checked()

        self._validate_dir_and_update_scan_button()

        self.select_all_btn.setEnabled(idle and has_rows)
        self.select_none_btn.setEnabled(idle and has_rows)
        self.delete_btn.setEnabled(idle and has_selection)
        self.download_btn.setEnabled(idle and has_rows)  # disabled when no results

        if idle:
            self._update_progress_summary()

    def _set_busy(self, busy: bool):
        self.browse_btn.setEnabled(not busy)
        self.dir_edit.setEnabled(not busy)
        self.filter_edit.setEnabled(not busy)
        self.clear_filter_btn.setEnabled(not busy)
        self.theme_combo.setEnabled(not busy)

        if busy:
            self.scan_btn.setEnabled(False)
            self.select_all_btn.setEnabled(False)
            self.select_none_btn.setEnabled(False)
            self.delete_btn.setEnabled(False)
            self.download_btn.setEnabled(False)
        else:
            self._refresh_action_states()

    # ---------- Sorting ----------

    def _apply_sort_indicator(self):
        self.grid.set_sort_indicator(self.sort_col, self.sort_ascending)

    def on_right_header_clicked(self, col: int):
        if self.progress_mode != self.MODE_IDLE:
            return
        if self.sort_col == col:
            self.sort_ascending = not self.sort_ascending
        else:
            self.sort_col = col
            self.sort_ascending = True

        self._apply_sort_indicator()
        self.apply_filter_and_refresh(force=False)
        self._save_settings()

    def _sort_rows(self, rows: List[FileRow]) -> List[FileRow]:
        reverse = not self.sort_ascending

        def key_fn(r: FileRow):
            if self.sort_col == self.SORT_DIRECTORY:
                return r.directory.lower()
            if self.sort_col == self.SORT_NAME:
                return r.name.lower()
            if self.sort_col == self.SORT_LENGTH:
                return r.length
            if self.sort_col == self.SORT_ATIME:
                return parse_dt(r.last_access_time) or datetime.min
            if self.sort_col == self.SORT_MTIME:
                return parse_dt(r.last_write_time) or datetime.min
            if self.sort_col == self.SORT_CTIME:
                return parse_dt(r.creation_time) or datetime.min
            return r.name.lower()

        return sorted(rows, key=key_fn, reverse=reverse)

    # ---------- Filtering / refresh ----------

    def apply_filter_and_refresh(self, force: bool = False):
        # block user-triggered refresh during scan/delete, but allow internal forced refresh
        if (self.progress_mode != self.MODE_IDLE) and (not force):
            return

        needle = (self.filter_edit.text() or "").strip().lower()

        if not needle:
            filtered = self.all_rows[:]
        else:
            filtered = [r for r in self.all_rows if needle in f"{r.directory} {r.name}".lower()]

        filtered = self._sort_rows(filtered)
        self._apply_sort_indicator()

        checked_before = set(self.grid.selected_paths())
        self.grid.populate(filtered, checked_before)

        total = len(self.all_rows)
        shown = len(filtered)
        if total == 0:
            self.status_label.setText("Ready.")
        elif needle:
            self.status_label.setText(f"Showing {shown} of {total} files (filtered).")
        else:
            self.status_label.setText(f"Showing {shown} files.")

        self._refresh_action_states()

    # ---------- Streaming flush ----------

    def _flush_stream_buffer(self):
        if not self._stream_buffer:
            return

        # During scan, we stream only when filter is empty (fast + intuitive).
        # If a filter is active, we wait for the end-of-scan full refresh.
        needle = (self.filter_edit.text() or "").strip().lower()
        if needle:
            self._stream_buffer.clear()
            return

        # Append buffered rows (unsorted streaming), then final sort happens at end.
        for r in self._stream_buffer:
            self.grid.append_row(r, is_checked=False)
        self._stream_buffer.clear()

        # keep buttons/progress summary up to date
        self._refresh_action_states()

    # ---------- Actions ----------

    def on_browse(self):
        if self.progress_mode != self.MODE_IDLE:
            return
        directory = QFileDialog.getExistingDirectory(self, "Select ComfyUI Directory")
        if directory:
            self.dir_edit.setText(directory)
            self._save_settings()

    def on_scan(self):
        if self.progress_mode != self.MODE_IDLE:
            return

        root_dir = self.dir_edit.text().strip()
        if not root_dir or not Path(root_dir).exists():
            QMessageBox.warning(self, "Missing Directory", "Please select a valid ComfyUI directory first.")
            return

        self._set_busy(True)
        self._set_progress_mode(self.MODE_SCANNING)
        self.status_label.setText("Starting scan...")

        self.all_rows = []
        self.grid.clear()
        self._stream_buffer.clear()
        self._stream_timer.start()

        self.scan_worker = ScanWorker(root_dir)
        self.scan_worker.progress.connect(self.progress.setValue)
        self.scan_worker.status.connect(self.status_label.setText)
        self.scan_worker.row_found.connect(self.on_scan_row_found)
        self.scan_worker.done.connect(self.on_scan_done)
        self.scan_worker.error.connect(self.on_worker_error)
        self.scan_worker.start()

    def on_scan_row_found(self, row: FileRow):
        self.all_rows.append(row)
        # buffer for UI (avoid per-row repaint)
        self._stream_buffer.append(row)

    def on_scan_done(self, total: int):
        self._stream_timer.stop()
        self._stream_buffer.clear()

        # final sorted/filtered render (restores sorting + arrows)
        self.apply_filter_and_refresh(force=True)

        self._set_busy(False)
        self._set_progress_mode(self.MODE_IDLE)
        self._save_settings()

    def on_select_all(self):
        if self.progress_mode != self.MODE_IDLE:
            return
        self.grid.left.blockSignals(True)
        for i in range(self.grid.left.rowCount()):
            it = self.grid.left.item(i, 0)
            if it:
                it.setCheckState(Qt.CheckState.Checked)
        self.grid.left.blockSignals(False)
        self._refresh_action_states()

    def on_select_none(self):
        if self.progress_mode != self.MODE_IDLE:
            return
        self.grid.left.blockSignals(True)
        for i in range(self.grid.left.rowCount()):
            it = self.grid.left.item(i, 0)
            if it:
                it.setCheckState(Qt.CheckState.Unchecked)
        self.grid.left.blockSignals(False)
        self._refresh_action_states()

    def on_delete(self):
        if self.progress_mode != self.MODE_IDLE:
            return

        paths = self.grid.selected_paths()
        if not paths:
            return

        mode = "Move to Recycle Bin" if self.recycle_chk.isChecked() else "PERMANENTLY delete"
        msg = f"{mode} {len(paths)} selected file(s)?"
        if QMessageBox.question(self, "Confirm Delete", msg) != QMessageBox.StandardButton.Yes:
            return

        self._set_busy(True)
        self._set_progress_mode(self.MODE_DELETING)
        self.status_label.setText("Starting deletion...")

        self.del_worker = DeleteWorker(paths, use_recycle_bin=self.recycle_chk.isChecked())
        self.del_worker.progress.connect(self.progress.setValue)
        self.del_worker.status.connect(self.status_label.setText)
        self.del_worker.deleted.connect(self.on_deleted)
        self.del_worker.failed.connect(self.on_delete_failed)
        self.del_worker.error.connect(self.on_worker_error)
        self.del_worker.finished.connect(self.on_delete_finished)
        self.del_worker.start()

    def on_deleted(self, deleted_paths: list):
        deleted_set = set(deleted_paths)
        self.all_rows = [r for r in self.all_rows if r.full_path not in deleted_set]
        self.apply_filter_and_refresh(force=True)

    def on_delete_finished(self):
        self._set_busy(False)
        self._set_progress_mode(self.MODE_IDLE)

    def on_delete_failed(self, failures: list):
        if not failures:
            return
        lines = "\n".join([f"- {p}\n  {e}" for (p, e) in failures[:10]])
        extra = "" if len(failures) <= 10 else f"\n\n(and {len(failures)-10} more...)"
        QMessageBox.warning(self, "Some files could not be deleted", lines + extra)

    def on_download(self):
        if self.progress_mode != self.MODE_IDLE:
            return
        if self.grid.row_count() == 0:
            return

        default_name = "comfyui_models.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel Spreadsheet",
            default_name,
            "Excel Workbook (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        # export what is currently visible in the grid (honors filter + sort)
        rows = self.grid.visible_rows_as_filerows()
        if not rows:
            return

        try:
            self._write_xlsx(path, rows)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", f"Failed to write Excel file:\n{e}")
            return

        # launch Excel if available, otherwise do nothing
        self._launch_excel_if_available(path)

    def _write_xlsx(self, file_path: str, rows: List[FileRow]):
        wb = Workbook()
        ws = wb.active
        ws.title = "Models"

        headers = ["Directory", "Name", "Length", "LastAccessTime", "LastWriteTime", "CreationTime"]
        ws.append(headers)

        for r in rows:
            ws.append([r.directory, r.name, r.length, r.last_access_time, r.last_write_time, r.creation_time])

        ws.freeze_panes = "A2"

        # Auto-size columns (FIXED syntax)
        for col_idx, header in enumerate(headers, start=1):
            c = get_column_letter(col_idx)
            max_len = len(header)
            for cell in ws[c]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[c].width = min(max_len + 2, 80)

        wb.save(file_path)
        self.status_label.setText("Exported. (Launching Excel if available...)")

    def _launch_excel_if_available(self, file_path: str):
        try:
            import win32com.client  # type: ignore
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True
            excel.Workbooks.Open(os.path.abspath(file_path))
            self.status_label.setText("Exported and opened in Excel.")
        except Exception:
            # per requirement: if Excel isn't available, do nothing extra
            self.status_label.setText("Exported.")

    def on_worker_error(self, msg: str):
        self._stream_timer.stop()
        self._set_busy(False)
        self._set_progress_mode(self.MODE_IDLE)
        QMessageBox.critical(self, "Error", msg)

    # ---------- state ----------

    def _set_progress_mode(self, mode: str):
        self.progress_mode = mode
        if mode == self.MODE_SCANNING:
            self.progress.setValue(0)
            self.progress.setFormat("Scanning… %p%")
        elif mode == self.MODE_DELETING:
            self.progress.setValue(0)
            self.progress.setFormat("Deleting… %p%")
        else:
            self._update_progress_summary()
        self._refresh_action_states()

    def _update_progress_summary(self):
        count, total_bytes = self.grid.selected_count_and_size()
        self.progress.setValue(100)
        self.progress.setFormat("Selected: 0 files" if count == 0 else f"Selected: {count} files, {fmt_bytes(total_bytes)}")
        self.progress.repaint()

    def _set_busy(self, busy: bool):
        self.browse_btn.setEnabled(not busy)
        self.dir_edit.setEnabled(not busy)
        self.filter_edit.setEnabled(not busy)
        self.clear_filter_btn.setEnabled(not busy)
        self.theme_combo.setEnabled(not busy)

        if busy:
            self.scan_btn.setEnabled(False)
            self.select_all_btn.setEnabled(False)
            self.select_none_btn.setEnabled(False)
            self.delete_btn.setEnabled(False)
            self.download_btn.setEnabled(False)
        else:
            self._refresh_action_states()

    def _validate_dir_and_update_scan_button(self):
        if self.progress_mode != self.MODE_IDLE:
            self.scan_btn.setEnabled(False)
            return
        p = self.dir_edit.text().strip()
        ok = bool(p) and Path(p).exists() and Path(p).is_dir()
        self.scan_btn.setEnabled(ok)

    def _refresh_action_states(self):
        idle = (self.progress_mode == self.MODE_IDLE)
        has_rows = self.grid.row_count() > 0
        has_selection = self.grid.any_checked()

        self._validate_dir_and_update_scan_button()

        self.select_all_btn.setEnabled(idle and has_rows)
        self.select_none_btn.setEnabled(idle and has_rows)
        self.delete_btn.setEnabled(idle and has_selection)
        self.download_btn.setEnabled(idle and has_rows)

        if idle:
            self._update_progress_summary()


def main():
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
